import os
import zipfile
import glob
import shutil
import pandas as pd
import csv
import openpyxl
from natsort import natsorted
from datetime import datetime
from sqlalchemy import create_engine

# ==================== 配置区 ====================
ROOT_FOLDER = r"C:\Users\Administrator\Desktop\聚水潭\2026-06-11"

# MySQL 连接信息
MYSQL_CONFIG = {
    "host": "xxxxxxxxxxx",
    "user": "xxxxxx",
    "password": "xxxxxx",
    "database": "xxxxxx",
    "port": 3306,
    "charset": "utf8mb4"
}

# 中英文字段映射（你提供的完整字典）
FIELD_MAPPING = {
    "内部订单号": "internal_order_no",
    "线上订单号": "online_order_no",
    "子订单号": "sub_order_no",
    "售后单号": "after_sale_no",
    "支付时间": "payment_time",
    "订单完成时间": "order_complete_time",
    "发货时间": "ship_time",
    "订单时间": "order_time",
    "店铺名称": "shop_name",
    "店铺编码": "shop_code",
    "商品编码": "product_code",
    "款式编码(参考)": "style_code",
    "店铺商品编码": "shop_product_code",
    "店铺款式编码": "shop_style_code",
    "链接　": "link",   # 注意中文含有全角空格
    "品牌": "brand",
    "分类": "category",
    "虚拟分类": "virtual_category",
    "颜色及规格": "color_spec",
    "商品名称": "product_name",
    "线下客户编号": "offline_customer_no",
    "线下客户名称": "offline_customer_name",
    "唯一码": "unique_code",
    "商品数量": "product_quantity",
    "商品金额": "product_amount",
    "商品成本": "product_cost",
    "其中：红冲订单成本": "red_charge_order_cost",
    "商家优惠": "merchant_discount",
    "运费收入": "freight_income",
    "付款金额": "payment_amount",
    "其中：平台优惠": "platform_discount",
    "其中：红冲订单金额": "red_charge_order_amount",
    "实发金额": "actual_amount",
    "实发数量": "actual_quantity",
    "实发成本": "actual_cost",
    "订单类型": "order_type",
    "登记时间": "register_time",
    "确认时间": "confirm_time",
    "进仓时间": "inbound_time",
    "线上商品名称": "online_product_name",
    "发货前申请数量": "pre_ship_apply_quantity",
    "发货前申请金额": "pre_ship_apply_amount",
    "发货前申请成本": "pre_ship_apply_cost",
    "发货前退回数量": "pre_ship_return_quantity",
    "发货前退回金额": "pre_ship_return_amount",
    "发货前退回成本": "pre_ship_return_cost",
    "发货前退款金额": "pre_ship_refund_amount",
    "发货前退款成本": "pre_ship_refund_cost",
    "发货前退平台补贴": "pre_ship_platform_subsidy_refund",
    "发货后申请数量": "post_ship_apply_quantity",
    "发货后申请金额": "post_ship_apply_amount",
    "发货后申请成本": "post_ship_apply_cost",
    "发货后退回数量": "post_ship_return_quantity",
    "发货后退回金额": "post_ship_return_amount",
    "发货后退回成本": "post_ship_return_cost",
    "发货后退款金额": "post_ship_refund_amount",
    "发货后退款成本": "post_ship_refund_cost",
    "发货后退平台补贴": "post_ship_platform_subsidy_refund",
    "其中:次品商品数量": "defective_quantity",
    "退款类型": "refund_type",
    "问题类型": "issue_type",
    "明细退货类型": "detail_return_type",
    "分类单": "classification_sheet",
    "赠品": "gift",
    "组合装": "combo",
    "订单标签": "order_tag",
    "售后单标签": "after_sale_tag",
    "商品标签": "product_tag",
    "发货仓编号": "ship_warehouse_no",
    "发货仓库": "ship_warehouse_name",
    "运营云仓": "operation_cloud_warehouse",
    "收货仓编号": "receive_warehouse_no",
    "收货仓": "receive_warehouse_name",
    "供销商编号": "supplier_no",
    "供销商": "supplier_name",
    "快递单号": "tracking_no",
    "快递公司": "carrier",
    "省": "province",
    "市": "city",
    "期间": "period",
    "业务员": "salesperson",
    "【商品资料】：商品标签": "product_info_tag",
    "【商品资料】：商品简称": "product_short_name",
    "【商品资料】：颜色及规格": "product_color_spec",
    "【商品资料】：国标码": "product_gb_code",
    "【商品资料】：辅助码": "product_aux_code",
    "【商品资料】：单位": "product_unit",
    "【商品资料】：主仓位": "product_main_location",
    "【商品资料】：补充仓位": "product_supplement_location",
    "【商品资料】：备注": "product_remark",
    "【商品资料】：供应商商品编码": "supplier_product_code",
    "【商品资料】：供应商商品款号": "supplier_product_model",
    "【商品资料】：供应商编号": "product_supplier_code",
    "【商品资料】：供应商名": "product_supplier_name",
    "【商品资料】：商品属性": "product_attribute",
    "【商品资料】：是否天猫直送": "is_tmall_direct",
    "【商品资料】：其它属性1": "other_attr1",
    "【商品资料】：其它属性2": "other_attr2",
    "【商品资料】：其它属性3": "other_attr3",
    "【商品资料】：其它属性4": "other_attr4",
    "【商品资料】：其它属性5": "other_attr5",
    "【商品资料】：基本售价": "base_price",
    "【商品资料】：市场|吊牌价": "market_price",
    "【商品资料】：运营采购成本": "operation_procurement_cost",
    "【商品资料】：其它价格2": "other_price2",
    "【商品资料】：其它价格3": "other_price3",
    "【商品资料】：其它价格4": "other_price4",
    "【商品资料】：其它价格5": "other_price5",
    "【商品资料】：长(cm)": "length_cm",
    "【商品资料】：宽(cm)": "width_cm",
    "【商品资料】：高(cm)": "height_cm",
    "【商品资料】：体积(cm3)": "volume_cm3",
    "【商品资料】：重量(kg)": "weight_kg",
    "成本方案": "cost_plan"
}

# 关键词 -> 目标表名（包含关系）
# 压缩包/CSV 文件名包含这些关键词（任一），就导入对应的表
# 注意顺序：更具体的关键词应放在前面，避免误匹配
KEYWORD_TABLE_MAPPING = [
    ("新媒体成本", "cost_new_media"),
    ("三网成本",   "cost_three_network"),
    ("公司成本",   "cost_company"),
    ("分销成本",   "cost_distribution"),
]

# 导入模式：'append' 追加，'replace' 替换（一般用追加）
IF_EXISTS = 'append'

# 手动指定本次导入的时间范围（运行前修改这里）
TIME_RANGE = "2026-06-11"

# MySQL 写入分块大小（行数）
DB_CHUNK_SIZE = 10000
# ================================================

def excel转csv并映射字段(excel_path, csv_path, field_mapping, append_data=False):
    """
    流式读取 Excel，将中文字段名映射为英文，写入 CSV
    append_data=False 时写入表头；True 时只追加数据（不写表头）
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # 获取表头（第一行）
    headers = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    if not headers:
        wb.close()
        return 0

    # 映射表头：中 -> 英，只保留映射中存在的列，并保持原顺序
    mapped_headers = []
    for ch in headers:
        if ch in field_mapping:
            mapped_headers.append(field_mapping[ch])
        # 不在映射中的列直接忽略（不写入 CSV）
    if not mapped_headers:
        wb.close()
        return 0

    # 打开 CSV 文件
    mode = 'a' if append_data else 'w'
    with open(csv_path, mode, encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        # 写入表头（仅当非追加模式）
        if not append_data:
            writer.writerow(mapped_headers)

        # 逐行读取数据行（从第2行开始）
        data_row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 根据映射后的列顺序，从原行中提取对应值
            # 注意：原行顺序与 headers 一致，mapped_headers 是原列的子集
            # 需要找到每个 mapped_header 在原列中的位置
            # 为了提高效率，先构建一个索引映射
            if data_row_count == 0:
                # 第一次遇到数据行时，构建原列索引 -> mapped_headers 索引的映射
                # 实际上可以提前构建，但为了清晰，在这里构建一次
                col_index_map = []
                for eng_col in mapped_headers:
                    # 找到原 headers 中哪个中文对应这个英文
                    for idx, ch in enumerate(headers):
                        if ch in field_mapping and field_mapping[ch] == eng_col:
                            col_index_map.append(idx)
                            break
                # 如果没有正确映射，可能会出错，但假定映射完整
            # 提取数据
            data_row = [row[idx] if idx < len(row) else None for idx in col_index_map]
            writer.writerow(data_row)
            data_row_count += 1

    wb.close()
    return data_row_count

def csv导入mysql(csv_path, target_table, engine, if_exists_mode):
    """将 CSV 文件分块导入 MySQL 表"""
    csv_reader = pd.read_csv(csv_path, chunksize=DB_CHUNK_SIZE, encoding='utf-8-sig')
    first_chunk = True
    now = datetime.now()
    for chunk in csv_reader:
        chunk['insert_time'] = now
        chunk['time_range'] = TIME_RANGE
        chunk.to_sql(target_table, engine,
                     if_exists=if_exists_mode if first_chunk else 'append',
                     index=False)
        print(f"    已写入 {len(chunk)} 行")
        first_chunk = False
    print(f"  💾 数据导入完成")


def 根据文件名获取表名(filename: str) -> str | None:
    """根据文件名包含的关键词返回目标表名，未匹配则返回 None"""
    for keyword, table_name in KEYWORD_TABLE_MAPPING:
        if keyword in filename:
            return table_name
    return None


def 处理压缩包(zip_path, root_output, engine, field_mapping, if_exists_mode):
    zip_name = os.path.splitext(os.path.basename(zip_path))[0]
    target_dir = os.path.join(root_output, zip_name)

    print(f"\n📦 处理压缩包: {zip_name}")

    # 1. 清理并解压
    if os.path.exists(target_dir):
        shutil.rmtree(target_dir)
    os.makedirs(target_dir)

    with zipfile.ZipFile(zip_path, 'r') as zf:
        zf.extractall(target_dir)
    print(f"  解压至: {target_dir}")

    # 2. 重命名 Excel 文件（按自然顺序）
    excel_files = glob.glob(os.path.join(target_dir, "*.xlsx"))
    if not excel_files:
        print(f"  ⚠️ 未找到 .xlsx 文件，跳过")
        return

    sorted_excel = natsorted(excel_files)
    for idx, old_path in enumerate(sorted_excel, start=1):
        new_name = f"{zip_name}_{idx}.xlsx"
        new_path = os.path.join(target_dir, new_name)
        os.rename(old_path, new_path)
        print(f"  重命名: {os.path.basename(old_path)} -> {new_name}")

    # 3. 合并所有 Excel 为 CSV（流式 + 映射）
    csv_path = os.path.join(target_dir, f"{zip_name}.csv")
    first_file = True
    total_rows = 0
    for idx in range(1, len(sorted_excel) + 1):
        file_path = os.path.join(target_dir, f"{zip_name}_{idx}.xlsx")
        if not os.path.exists(file_path):
            continue
        print(f"  合并: {os.path.basename(file_path)}")
        rows = excel转csv并映射字段(file_path, csv_path, field_mapping, append_data=not first_file)
        total_rows += rows
        first_file = False
    print(f"  ✅ CSV 生成: {csv_path} (共 {total_rows} 行数据)")

    # 4. 导入 MySQL — 根据 zip 文件名自动匹配目标表
    target_table = 根据文件名获取表名(zip_name)
    if target_table is None:
        raise ValueError(f"文件名「{zip_name}」未匹配到任何目标表，请在 KEYWORD_TABLE_MAPPING 中添加映射")
    print(f"  导入数据库表: {target_table}")
    csv导入mysql(csv_path, target_table, engine, if_exists_mode)

def main():
    db_url = f"mysql+pymysql://{MYSQL_CONFIG['user']}:{MYSQL_CONFIG['password']}@{MYSQL_CONFIG['host']}:{MYSQL_CONFIG['port']}/{MYSQL_CONFIG['database']}?charset={MYSQL_CONFIG['charset']}"
    engine = create_engine(db_url)

    zip_files = glob.glob(os.path.join(ROOT_FOLDER, "*.zip"))
    if not zip_files:
        print(f"❌ 在 {ROOT_FOLDER} 中未找到任何 .zip 文件")
        return

    print(f"📁 找到 {len(zip_files)} 个压缩包，开始处理...")
    for zip_path in zip_files:
        try:
            处理压缩包(zip_path, ROOT_FOLDER, engine, FIELD_MAPPING, IF_EXISTS)
        except Exception as e:
            print(f"  ❌ 处理失败: {e}")
            import traceback
            traceback.print_exc()
            continue

    print(f"\n🎉 全部处理完成！")

if __name__ == "__main__":
    main()
